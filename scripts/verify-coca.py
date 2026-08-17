import re
import sys
from pathlib import Path

from openpyxl import load_workbook


project_root = Path(__file__).resolve().parents[1]
lexicon_text = (project_root / "scripts" / "lexicon.mjs").read_text(encoding="utf-8")
selected_words = re.findall(r"w: '([^']+)'", lexicon_text)

workbook_path = project_root.parent / "COCA词频单词表.xlsx"
workbook = load_workbook(workbook_path, read_only=True, data_only=True)
sheet = workbook["1 lemmas"]
coca_words = {
    str(row[1]).lower()
    for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True)
    if row[1]
}

missing = [word for word in selected_words if word not in coca_words]
print({"selected": len(selected_words), "missing": missing})
sys.exit(1 if missing else 0)
